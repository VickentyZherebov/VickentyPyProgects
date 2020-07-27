from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_workbook

import os

from ExportCandidates import export_candidates, count_max_comments
from LoadCandidates import load_candidates, load_history, load_comments

workbook = load_workbook('AIHUB.xlsx')
candidates = load_candidates(workbook)
load_history(workbook, candidates)
load_comments(workbook, candidates)

result_workbook = Workbook()
export_candidates(result_workbook, candidates)
save_workbook(result_workbook, "output.xlsx")

for local_id, candidate in list(candidates.items())[:100]:
    print(candidate.status)

os.system("output.xlsx")
